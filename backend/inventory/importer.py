"""CSV / XLSX import for the parts master. Pure-python (csv + openpyxl)."""
import csv, io
from db import get_connection

FIELDS = ["part_number", "description", "vendor", "vendor_name", "lead_time_weeks",
          "abc_group", "on_hand", "on_order", "forecast", "last_year_usage",
          "unit_price", "product_per_case", "cases_per_pallet", "eoq", "safety_stock"]
INT = {"lead_time_weeks", "on_hand", "on_order", "product_per_case", "cases_per_pallet", "eoq", "safety_stock"}
FLOAT = {"forecast", "last_year_usage", "unit_price"}


def _coerce(row: dict) -> tuple:
    out = []
    for f in FIELDS:
        v = row.get(f, "")
        if f in INT:
            v = int(float(v or 0))
        elif f in FLOAT:
            v = float(v or 0)
        out.append(v)
    return tuple(out)


def parse_csv(data: bytes) -> list[dict]:
    return list(csv.DictReader(io.StringIO(data.decode("utf-8-sig"))))


def parse_xlsx(data: bytes) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() for h in next(rows)]
    return [dict(zip(header, r)) for r in rows if any(c is not None for c in r)]


def import_parts(rows: list[dict]) -> int:
    missing = [r["part_number"] for r in rows if not r.get("part_number")]
    if missing:
        raise ValueError("All rows must have part_number")
    conn = get_connection()
    cur = conn.cursor()
    ph = ",".join(["?"] * len(FIELDS))
    cur.executemany(f"INSERT OR REPLACE INTO parts ({','.join(FIELDS)}) VALUES ({ph})",
                    [_coerce(r) for r in rows])
    conn.commit(); cur.close(); conn.close()
    return len(rows)
